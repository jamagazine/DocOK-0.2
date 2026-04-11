import os
import re
import pandas as pd
import io
import pdfplumber
import cv2
import numpy as np
from PIL import Image
from thefuzz import fuzz

CLIENT_INN = "5904003027"  # ПАО ММК-Пермь
CLIENT_ADDRESS = "614058, ПЕРМСКИЙ КРАЙ, ПЕРМЬ Г, ПРОМЫШЛЕННАЯ УЛ, ДОМ 110"

def clean_val(x):
    if pd.isna(x) or x is None:
        return ""
    s = str(x).strip()
    if s.lower() in ['nan', 'none', 'null']:
        return ""
    # Очистка .0 для числовых строк (устойчиво к пробелам)
    if s.endswith('.0') and s.replace('.0', '').replace(' ', '').isdigit():
        s = s.split('.')[0].strip()
    
    # Удаляем лишние кавычки из начала и конца, если они есть
    if s.startswith('"') and s.endswith('"'):
        s = s[1:-1]
    # Экранируем пайпы для корректного Markdown
    return s.replace('|', '\\|').replace('\n', ' ')

def deskew_image(pil_img):
    """
    Straightens a tilted scan (OCR Stage 1 Optimization).
    Uses OpenCV to find the text angle and rotate the image.
    """
    # 1. Convert PIL to OpenCV (BGR)
    cv_img = cv2.cvtColor(np.array(pil_img), cv2.COLOR_RGB2BGR)
    
    # 2. To Grayscale and Binarize
    gray = cv2.cvtColor(cv_img, cv2.COLOR_BGR2GRAY)
    gray = cv2.bitwise_not(gray)
    thresh = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY | cv2.THRESH_OTSU)[1]
    
    # 3. Find coordinates of all white pixels (text)
    coords = np.column_stack(np.where(thresh > 0))
    if len(coords) == 0: return pil_img

    # 4. Find the minAreaRect for the points
    rect = cv2.minAreaRect(coords)
    angle = rect[-1]
    
    # Normalize the angle (OpenCV returns -90 to 0)
    if angle < -45:
        angle = -(90 + angle)
    else:
        angle = -angle
        
    # Rotate if tilt is significant (> 0.5 deg)
    if abs(angle) > 0.5:
        (h, w) = cv_img.shape[:2]
        center = (w // 2, h // 2)
        M = cv2.getRotationMatrix2D(center, angle, 1.0)
        cv_img = cv2.warpAffine(cv_img, M, (w, h), flags=cv2.INTER_CUBIC, borderMode=cv2.BORDER_REPLICATE)
        # Convert back to PIL
        return Image.fromarray(cv2.cvtColor(cv_img, cv2.COLOR_BGR2RGB))
    
    return pil_img

UNITS_MAP = {
    # базовые
    "шт": "штука", "шт.": "штука", "штук": "штука", "штуки": "штука",
    "м": "метр", "м.": "метр", "метр": "метр", "метры": "метр",
    "м2": "квадратный метр", "м²": "квадратный метр", "м2.": "квадратный метр",
    "м3": "кубический метр", "м³": "кубический метр",
    "кг": "килограмм", "кг.": "килограмм",
    "т": "тонна", "т.": "тонна", "тн": "тонна", "тн.": "тонна",
    "л": "литр", "л.": "литр",
    "компл": "комплект", "компл.": "комплект",
    "наб": "набор", "наб.": "набор",
    "рул": "рулон", "рул.": "рулон",
    "упак": "упаковка", "упак.": "упаковка", "пач": "пачка", "пач.": "пачка",
    "пог.м": "погонный метр", "п.м": "погонный метр", "погм": "погонный метр",
    "тыс.шт": "тысяча штук", "тыс.шт.": "тысяча штук",
    "mm": "миллиметр", "мм": "миллиметр", "мм.": "миллиметр",
    "cm": "сантиметр", "см": "сантиметр", "см.": "сантиметр",
    "km": "километр", "км": "километр", "км.": "километр",
    "ga": "гектар",
    "кв.м": "квадратный метр", "кв.м.": "квадратный метр",
    "куб.м": "кубический метр", "куб.м.": "кубический метр",
    # реже, но встречаются в стройке
    "пар": "пара", "пар.": "пара",
    "мест": "место", "мест.": "место",
    "секц": "секция", "секц.": "секция",
    "эл": "элемент", "эл.": "элемент",
    "бух": "бухта", "бух.": "бухта",
    "лист": "лист", "лист.": "лист",
    "кг/м": "килограмм на метр",
    "м/пог": "метр погонный",
    "кор": "коробка", "кор.": "коробка", "ящ": "ящик", "ящ.": "ящик",
    "банка": "банка", "б.": "банка"
}

def normalize_unit(unit_str: str) -> str:
    if not unit_str:
        return ""
    cleaned = unit_str.lower().strip()
    return UNITS_MAP.get(cleaned, unit_str)

def normalize_for_match(text: str) -> str:
    if not text: return ""
    t = text.lower().strip()
    t = re.sub(r'[^\w\sа-яё]', ' ', t)
    t = re.sub(r'\s+', ' ', t).strip()
    return t

def to_float(val) -> float:
    if not val: return 0.0
    val_str = str(val).replace(' ', '').replace(',', '.')
    match = re.search(r'-?\d+(\.\d+)?', val_str)
    return float(match.group(0)) if match else 0.0

def validate_and_clean_inn(text: str) -> str:
    """
    Validates Russian INN (10 or 12 digits).
    Includes logic for 11-digit OCR errors (stripping first/last char).
    """
    import re
    inn = re.sub(r'\D', '', str(text))
    
    # 1. Handle OCR ghosts (11 digits)
    if len(inn) == 11:
        candidates = [inn[1:], inn[:-1]]
        for cand in candidates:
            if validate_inn_logic(cand): return cand
    
    return inn if validate_inn_logic(inn) else ""

def validate_inn_logic(s: str) -> bool:
    """
    Strict Russian INN checksum validation (10 or 12 digits).
    """
    if not s or not s.isdigit(): return False
    if len(s) not in [10, 12]: return False
    
    if len(s) == 10:
        coeffs = [2, 4, 10, 3, 5, 9, 4, 6, 8]
        s_sum = sum(int(s[i]) * coeffs[i] for i in range(9))
        return (s_sum % 11) % 10 == int(s[9])
        
    if len(s) == 12:
        coeffs1 = [7, 2, 4, 10, 3, 5, 9, 4, 6, 8]
        s_sum1 = sum(int(s[i]) * coeffs1[i] for i in range(10))
        n11 = (s_sum1 % 11) % 10
        
        coeffs2 = [3, 7, 2, 4, 10, 3, 5, 9, 4, 6, 8]
        s_sum2 = sum(int(s[i]) * coeffs2[i] for i in range(11))
        n12 = (s_sum2 % 11) % 10
        
        return n11 == int(s[10]) and n12 == int(s[11])
    return False

def validate_and_clean_inn(text: str) -> str:
    import re
    inn = re.sub(r'\D', '', str(text))
    if len(inn) == 11:
        for cand in [inn[1:], inn[:-1]]:
            if validate_inn_logic(cand): return cand
    return inn if validate_inn_logic(inn) else ""

def validate_bank_details(bik: str, acc_settlement: str = "", acc_corr: str = ""):
    import re
    bik = re.sub(r'\D', '', str(bik))
    acc_s = re.sub(r'\D', '', str(acc_settlement))
    acc_c = re.sub(r'\D', '', str(acc_corr))
    is_bik_ok = len(bik) == 9 and bik.startswith('04')
    res_s = acc_s if len(acc_s) == 20 else ""
    res_c = acc_c if len(acc_c) == 20 and acc_c.startswith('301') else ""
    is_valid = is_bik_ok and res_c and bik[-3:] == res_c[-3:]
    return {"bik": bik if is_bik_ok else "", "settlement_account": res_s, 
            "correspondent_account": res_c, "is_valid": is_valid}

def calculate_uncertainty(struct: dict, global_low_conf: bool):
    items = struct.get("items", [])
    for item in items:
        is_uncertain = global_low_conf
        raw_qty = str(item.get("quantity", "")).strip()
        match_qty = re.match(r'^([\d\.\,\s]+)(.*?)$', raw_qty)
        if match_qty and match_qty.group(2).strip():
            item["quantity"] = match_qty.group(1).strip()
            if not item.get("unit"): item["unit"] = match_qty.group(2).strip()
                
        if item.get("unit"): item["unit"] = normalize_unit(str(item.get("unit")))

        qty = to_float(item.get("quantity"))
        price = to_float(item.get("price"))
        total = to_float(item.get("total"))
        
        if qty > 0 and price > 0 and total > 0:
            if abs(qty * price - total) > (0.05 * total): is_uncertain = True
            if abs(qty * price - total) >= 0.1: item["math_error"] = True
                
        if not item.get("name"): is_uncertain = True
        item["isUncertain"] = is_uncertain
        
    return struct

def transliterate(text: str) -> str:
    ru = "абвгдёезийклмнопрстуфхцчшщъыьэюяАБВГДЁЕЗИЙКЛМНОПРСТУФХЦЧШЩЪЫЬЭЮЯ"
    en = ["a", "b", "v", "g", "d", "yo", "e", "z", "i", "j", "k", "l", "m", "n", "o", "p", "r", "s", "t", "u", "f", "h", "ts", "ch", "sh", "shch", "", "y", "", "e", "yu", "ya", "A", "B", "V", "G", "D", "Yo", "E", "Z", "I", "J", "K", "L", "M", "N", "O", "P", "R", "S", "T", "U", "F", "H", "Ts", "Ch", "Sh", "Shch", "", "Y", "", "E", "Yu", "Ya"]
    return "".join({ru[i]: en[i] for i in range(len(ru))}.get(c, c) for c in text)

def secure_filename(filename: str) -> str:
    return re.sub(r'[^a-zA-Z0-9._-]', '_', filename)

def sanitize_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    df = df.fillna("").astype(str)
    c_p, c_n = 0, 1
    # Fix Column Mapping: Added break to fix first match
    for i, c in enumerate(df.columns):
        cl = str(c).lower()
        if any(x in cl for x in ["поз", "№", "unnamed: 0"]): 
            c_p = i
            break
    for i, c in enumerate(df.columns):
        cl = str(c).lower()
        if any(x in cl for x in ["наименован", "названи", "товар"]):
            c_n = i
            break

    last = ""
    for i in range(len(df)):
        p, n = df.iloc[i, c_p].strip(), df.iloc[i, c_n].strip()
        
        # Merged Cells Support: Check neighbor column if c_n is empty
        if not n and c_n != 1:
            n_alt = df.iloc[i, 1].strip()
            if n_alt:
                n = n_alt
                df.iloc[i, c_n] = n
        
        if (not p or p=="nan") and n:
            # Regex Update: Use \s* instead of \s+ for missing spaces
            m = re.match(r'^(\d+(?:\.\d+)*)\.?\s*(.*)', n)
            if m:
                df.iloc[i, c_p], df.iloc[i, c_n], p = m.group(1), m.group(2), m.group(1)
        
        if p.endswith(".1") and last.endswith(".9") and p[:-2]==last[:-2]: df.iloc[i, c_p] = p + "0"
        if re.match(r'^\d+(\.\d+)+$', p): last = p
        
    # Append technical ID column for AI patching
    df['__ID__'] = [f"idx_{i}" for i in range(len(df))]
    return df

def convert_df_to_items(df: pd.DataFrame) -> list:
    c_p, c_n, c_u, c_q, c_id = 0, 1, -1, -1, -1
    for i, c in enumerate(df.columns):
        cl = str(c).lower()
        if "поз" in cl or "№" in cl: c_p = i
        elif "наимен" in cl or "товар" in cl: c_n = i
        elif "ед" in cl and "изм" in cl: c_u = i
        elif "кол" in cl: c_q = i
        elif "__id__" in cl: c_id = i

    def is_empty_val(x):
        return str(x).strip().lower() in ("", "0", "0.0", "nan", "none")

    # ── Step 1: Classify every row as HEADER or ITEM ──────────────────────────
    # ITEM  = has a readable quantity
    # HEADER= name present, quantity absent
    raw_list = []
    for idx, row in df.iterrows():
        v = [str(x).strip() for x in row.values]
        if not any(v): continue
        # Skip the column-header row of the table itself (1, 2, 3, ...)
        if v[c_p] == "1" and v[c_n] == "2" and sum(1 for i, x in enumerate(v[:5]) if x == str(i+1)) >= 3:
            continue

        row_id   = v[c_id] if c_id != -1 else f"idx_{idx}"
        name_val = v[c_n]
        pos_val  = v[c_p]
        unit_val = v[c_u] if c_u != -1 else ""
        qty_val  = v[c_q] if c_q != -1 else ""

        name_clean = str(name_val).strip()
        # Step 1: Preliminary identification of potential headers
        # Potential header = No unit, no quantity, no mass (columns after c_q) + HAS NAME WITH LETTERS
        has_letters = any(c.isalpha() for c in str(name_val))
        is_potential = is_empty_val(unit_val) and is_empty_val(qty_val) and bool(str(name_val).strip()) and has_letters

        raw_list.append({
            "id":       row_id,
            "pos":      pos_val,
            "name":     name_clean,
            "unit":     unit_val,
            "quantity": qty_val,
            "is_item":  not is_potential,  # True = ITEM, False = potential HEADER
        })

    # ── Step 2: Split into Chunks [headers…, items…] ──────────────────────────
    # One Chunk = 1+ consecutive HEADER rows + 1+ consecutive ITEM rows.
    chunks = []
    cur_headers: list = []
    cur_items:   list = []

    for raw in raw_list:
        if raw["is_item"]:
            cur_items.append(raw)
        else:
            # New header block starts — flush accumulated items first
            if cur_items:
                chunks.append({"headers": cur_headers, "items": cur_items})
                cur_headers = []
                cur_items   = []
            cur_headers.append(raw)

    # Flush the final chunk
    if cur_headers or cur_items:
        chunks.append({"headers": cur_headers, "items": cur_items})

    # ── Step 3: Stack AST – build hierarchy ───────────────────────────────────
    # current_stack[idx] is an ancestor header; idx == hierarchy level (0,1,2).
    LEVEL_TYPES   = ["WORK_TYPE", "LOCATION", "GROUP"]
    current_stack: list = []
    items_out:     list = []
    emitted_ids:   set  = set()

    def is_global_section(name: str) -> bool:
        """Single ALL-CAPS word with letters (e.g. ВЕНТИЛЯЦИЯ, OTOPLENIE)."""
        parts = name.split()
        return (
            len(parts) == 1
            and len(name) > 2
            and all(ch.isupper() or not ch.isalpha() for ch in name)
            and any(ch.isalpha() for ch in name)
        )

    for chunk in chunks:
        hdrs = chunk["headers"]
        its  = chunk["items"]

        if not hdrs and not its:
            continue

        # ── Update the stack based on this chunk's headers ────────────────────
        if hdrs:
            if len(hdrs) >= 2:
                # Multiple headers → completely new branch, reset stack
                current_stack = list(hdrs)
            else:
                hdr    = hdrs[0]
                name_h = hdr["name"]
                if is_global_section(name_h):
                    # Global ALL-CAPS word → new L0 root, reset stack
                    current_stack = [hdr]
                else:
                    # Inherit: replace only the deepest slot in the stack
                    if current_stack:
                        current_stack[-1] = hdr
                    else:
                        current_stack = [hdr]

        # ── Emit header rows with assigned levels ─────────────────────────────
        for depth, hdr in enumerate(current_stack):
            if hdr["id"] in emitted_ids:
                continue  # already emitted in a previous chunk
            emitted_ids.add(hdr["id"])

            lvl    = min(depth, 2)
            r_type = LEVEL_TYPES[lvl]

            if lvl == 0:
                hdr["pos"] = ""
            elif lvl == 1:
                hdr["pos"]  = "§"
                hdr["name"] = re.sub(r'^[§\s]+|[§\s]+$', '', hdr["name"]).strip()

            parent_id = current_stack[depth - 1]["id"] if depth > 0 else None

            hdr["row_type"]        = r_type
            hdr["is_header"]       = True
            hdr["hierarchy_level"] = lvl
            hdr["parentId"]        = parent_id
            hdr["note"]            = ""
            hdr.pop("is_item", None)
            items_out.append(hdr)

        # Reset so the same chunk headers aren't emitted again on the next loop
        current_stack = list(current_stack)

        # ── Emit item rows, parented to the deepest header ────────────────────
        deepest_parent = current_stack[-1]["id"] if current_stack else None
        for item in its:
            item["row_type"]        = "ITEM"
            item["is_header"]       = False
            item["hierarchy_level"] = None
            item["parentId"]        = deepest_parent
            item["note"]            = ""
            item.pop("is_item", None)
            items_out.append(item)

    return items_out

def extract_text_from_pdf(path: str) -> str:
    ext_text = ""
    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            t = page.extract_text(x_tolerance=2, y_tolerance=3)
            if t:
                ext_text += "\n".join([re.sub(r'\s{2,}', ' | ', l) for l in t.split('\n')]) + "\n"
    return ext_text

def extract_specification_summary(df: pd.DataFrame, parsed_rows: list, file_path: str = "") -> dict:
    """
    Extracts metadata from the GOST stamp using xlrd merged cells.
    The stamp data (Cipher, Destination) is stored in merged cells on the
    'Спецификация' sheet, which pandas does not read by default.
    Statistics (positions, suppliers) come from the parsed table data.
    """
    # ── 1. STATISTICS from parsed_rows (ITEMs only) ──────────────────────────
    items = [r for r in (parsed_rows or []) if r.get("row_type") == "ITEM"]
    total_positions = len(items)

    # Suppliers: Find the supplier column dynamically
    c_s = -1
    for i, c in enumerate(df.columns):
        cl = str(c).lower()
        if "поставщик" in cl or "изготовитель" in cl:
            c_s = i
            break
    
    if c_s == -1 and 4 < len(df.columns):
        c_s = 4 # Fallback to column 4 if not found but exists

    suppliers_map = {} # Case-insensitive map to preserve one of the casings
    for r in items:
        rid = r.get("id", "")
        if rid.startswith("idx_"):
            try:
                row_idx = int(rid[4:])
                if row_idx < len(df) and c_s != -1:
                    val = str(df.iloc[row_idx, c_s]).strip()
                    if val and val.lower() not in ("", "0", "0.0", "nan", "none"):
                        suppliers_map[val.lower()] = val
            except (ValueError, IndexError):
                pass
    suppliers_str = ", ".join(sorted(suppliers_map.values())) if suppliers_map else "Не определено"

    # ── 2. STAMP EXTRACTION via xlrd (merged cells) ──────────────────────────
    cipher = ""
    destination_parts = []
    debug_lines = []
    anchor_note = ""

    if file_path and file_path.lower().endswith(".xls"):
        try:
            import xlrd
            wb = xlrd.open_workbook(file_path, formatting_info=True)

            # Try to find the sheet with the stamp (prefer "Спецификация", fallback to any)
            stamp_sheet = None
            for sn in wb.sheet_names():
                if "специф" in sn.lower():
                    stamp_sheet = wb.sheet_by_name(sn)
                    break
            if not stamp_sheet and wb.nsheets > 1:
                stamp_sheet = wb.sheet_by_index(1)  # try second sheet
            if not stamp_sheet:
                stamp_sheet = wb.sheet_by_index(0)

            debug_lines.append(f"Sheet: '{stamp_sheet.name}', Rows: {stamp_sheet.nrows}, Merged: {len(stamp_sheet.merged_cells)}")

            # Scan merged cells for stamp data
            # Pattern: Cipher is in merged cells spanning columns 13-19 in the first 60 rows
            # Destination is in adjacent merged cells below the cipher
            stamp_merges = []
            for mc in stamp_sheet.merged_cells:
                r_lo, r_hi, c_lo, c_hi = mc
                if r_lo < 60:  # Only first page
                    val = stamp_sheet.cell_value(r_lo, c_lo)
                    if val:
                        val_str = str(val).strip()
                        stamp_merges.append({
                            "r_lo": r_lo, "r_hi": r_hi - 1,
                            "c_lo": c_lo, "c_hi": c_hi - 1,
                            "val": val_str
                        })

            # Sort by row for readability
            stamp_merges.sort(key=lambda x: (x["r_lo"], x["c_lo"]))

            # Visual Grid: Recreate the document slice around the anchor
            anchor_idx = -1
            for i in range(stamp_sheet.nrows):
                for j in range(stamp_sheet.ncols):
                    if "листов" in str(stamp_sheet.cell_value(i, j)).lower():
                        anchor_idx = i
                        break
                if anchor_idx != -1: break

            if anchor_idx != -1:
                debug_lines.append("")
                debug_lines.append("=== VISUAL GRID AROUND 'ЛИСТОВ' (STAMP AREA) ===")
                slice_top = max(0, anchor_idx - 15)
                slice_bot = min(stamp_sheet.nrows, anchor_idx + 5)
                for ri in range(slice_top, slice_bot):
                    row_cells = []
                    for ci in range(stamp_sheet.ncols):
                        # Ensure we see at least up to Column 20
                        if ci > 20: continue
                        val = str(stamp_sheet.cell_value(ri, ci)).strip()
                        if val == "nan": val = ""
                        row_cells.append(val or " ")
                    debug_lines.append(f"R{ri:03d} | {' | '.join(row_cells)} |")

            # ── CIPHER: Look for merged cells in columns 13+ with cipher pattern ──
            cipher_regex = r'(\d{2,}-\d{2,}-[А-Яа-яA-Za-z\.\d/_-]+)'
            for sm in stamp_merges:
                if sm["c_lo"] >= 13 and sm["r_lo"] < 60:
                    m = re.search(cipher_regex, sm["val"])
                    if m:
                        found = m.group(0)
                        if len(found) > len(cipher):
                            cipher = found

            # ── DESTINATION: Collect text from merged cells near cipher ──
            # Look for large text blocks in columns 13+ that are NOT the cipher
            noise_words = ["спецификац", "изм.", "лист", "стади", "архив", "инв.",
                           "формат", "дата", "подп", "разраб", "пров.", "н.контр",
                           "утв.", "копировал", "взам.", "№ док", "поз.", "наименован",
                           "код продукц", "поставщик", "ед. измер", "масса", "примечан",
                           "кол.", "взамен", "инв. №", "арх. №", "подпись", "дата", "инв.№"]
            for sm in stamp_merges:
                if sm["c_lo"] >= 13 and sm["r_lo"] < 60:
                    val = sm["val"]
                    if val == cipher:
                        continue
                    if len(val) < 6:
                        continue
                    if re.match(r'^[\d\.\,\s]+$', val):
                        continue
                    if any(nw in val.lower() for nw in noise_words):
                        continue
                    if val not in destination_parts:
                        destination_parts.append(val)

        except Exception as e:
            anchor_note = f"⚠ Ошибка чтения xlrd: {e}\n"
            debug_lines.append(anchor_note)

    elif file_path and file_path.lower().endswith(".xlsx"):
        # For .xlsx files, try openpyxl merged cells
        try:
            import openpyxl
            wbx = openpyxl.load_workbook(file_path, data_only=True)

            stamp_ws = None
            for sn in wbx.sheetnames:
                if "специф" in sn.lower():
                    stamp_ws = wbx[sn]
                    break
            if not stamp_ws and len(wbx.sheetnames) > 1:
                stamp_ws = wbx.worksheets[1]
            if not stamp_ws:
                stamp_ws = wbx.active

            debug_lines.append(f"Sheet: '{stamp_ws.title}', Merged: {len(stamp_ws.merged_cells.ranges)}")

            # Read merged cell ranges
            cipher_regex = r'(\d{2,}-\d{2,}-[А-Яа-яA-Za-z\.\d/_-]+)'
            for mr in stamp_ws.merged_cells.ranges:
                r_lo = mr.min_row - 1  # Convert to 0-indexed
                c_lo = mr.min_col - 1
                if r_lo < 60:
                    cell = stamp_ws.cell(mr.min_row, mr.min_col)
                    if cell.value:
                        val = str(cell.value).strip()
                        if c_lo >= 13:
                            m = re.search(cipher_regex, val)
                            if m and len(m.group(0)) > len(cipher):
                                cipher = m.group(0)
                            elif len(val) >= 6 and val != cipher:
                                if not re.match(r'^[\d\.\,\s]+$', val):
                                    if val not in destination_parts:
                                        destination_parts.append(val)
                        debug_lines.append(f"  R{r_lo:03d} C{c_lo:02d}: '{val[:70]}'")
        except Exception as e:
            anchor_note = f"⚠ Ошибка чтения openpyxl: {e}\n"
            debug_lines.append(anchor_note)

    else:
        # Fallback: no file path or unsupported format
        anchor_note = "⚠ Путь к файлу не указан или формат не поддерживается.\n"
        debug_lines.append(anchor_note)

        # Try anchor-based search in the pandas df as last resort
        anchor_idx = -1
        for i in range(min(len(df), 100)):
            for col_idx in range(len(df.columns)):
                cell = str(df.iloc[i, col_idx]).strip().lower()
                if "листов" in cell or "изм." in cell:
                    anchor_idx = i
                    break
            if anchor_idx != -1:
                break

        if anchor_idx != -1:
            slice_top = max(0, anchor_idx - 10)
            slice_bot = min(len(df), anchor_idx + 4)
            for ri in range(slice_top, slice_bot):
                cells = " | ".join(str(v).strip()[:40] for v in df.iloc[ri].values)
                debug_lines.append(f"R{ri:03d} | {cells} |")

    # --- Step 4: Extract general notes for the whole specification ---
    notes_parts = []
    extraction_log = "Init"
    if file_path and file_path.lower().endswith(".xls"):
        try:
            import xlrd
            wb_notes = xlrd.open_workbook(file_path, formatting_info=False)
            target_sheet = None
            for s in wb_notes.sheets():
                if "специф" in s.name.lower():
                    target_sheet = s
                    break
            if not target_sheet:
                target_sheet = wb_notes.sheet_by_index(min(1, wb_notes.nsheets - 1))

            if target_sheet:
                # 1. Identify "The Stamp Envelope" vertically
                # We use the row where the Cipher or Stamp Anchor was found
                stamp_top = -1
                stamp_bot = -1
                
                for r in range(target_sheet.nrows):
                    row_vals = [str(target_sheet.cell_value(r, c)).strip().lower() for c in range(target_sheet.ncols)]
                    row_str = " ".join(row_vals)
                    
                    # Top boundary: The Project Cipher or Name usually starts the stamp
                    if cipher and cipher.lower() in row_str:
                        if stamp_top == -1: stamp_top = r
                    
                    # Bottom boundary: Anchor "Листов" usually signals the end
                    if "листов" in row_str:
                        stamp_bot = r
                        if stamp_top == -1: stamp_top = max(0, r - 6) # Fallback if cipher not found
                        break
                
                if stamp_top != -1 and stamp_bot != -1:
                    extraction_log = f"Stamp range: R{stamp_top} to R{stamp_bot}"
                    # 2. Collect everything to the LEFT of signatures ONLY inside this envelope
                    # Scan from Cipher down to Stamp Bottom
                    for ri in range(stamp_top, stamp_bot + 2):
                        if ri >= target_sheet.nrows: break
                        row_vals = [str(target_sheet.cell_value(ri, ci)).strip() for ci in range(target_sheet.ncols)]
                        # Ignore columns where signatures usually reside (I-M)
                        left_zone = row_vals[:8]
                        for val in left_zone:
                            if len(val) > 15 or "примечание" in val.lower():
                                if val not in notes_parts:
                                    notes_parts.append(val)
                else:
                    extraction_log = f"Could not define stamp envelope (Top:{stamp_top}, Bot:{stamp_bot})"
        except Exception as e:
            extraction_log = f"Notes Error: {e}"

    # Concat into one sentence
    notes_str = " ".join(notes_parts) if notes_parts else "Отсутствуют"
    notes_str = re.sub(r'\s+', ' ', notes_str).strip()

    # ── 5. ASSEMBLE ──────────────────────────────────────────────────────────
    final_cipher = cipher or "В штампе не найден"
    final_dest = " — ".join(destination_parts) if destination_parts else "В штампе не найден"
    debug_grid = "\n".join(debug_lines)

    summary_md = f"### Общая сводка\n\n" \
                 f"**Номер спецификации:** {final_cipher}\n\n" \
                 f"**Назначение:** {final_dest}\n\n" \
                 f"**Позиций всего (объединённых):** {total_positions} шт.\n\n" \
                 f"**Поставщики в документе:** {suppliers_str}\n\n" \
                 f"**Примечания:**\n{notes_str}"

    return {
        "summary_md": summary_md,
        "debug_grid": debug_grid,
        "fields": {
            "cipher": final_cipher,
            "destination": final_dest,
            "total_positions": total_positions,
            "suppliers": suppliers_str,
            "notes": notes_str
        }
    }



def pdf_to_grid_markdown(file_path: str) -> str:
    """
    Converts a digital PDF invoice to Markdown using the Unified Zonal Pipeline.
    """
    try:
        pages_data = extract_digital_words_as_ocr(file_path)
        
        # 1. Заголовок (Зональный) - берем только первую страницу, чтобы избежать наслоения координат
        header_md = clean_and_build_markdown(pages_data[0] if pages_data else [])
        
        # 2. Табличная часть (Grid) - по каждой странице отдельно (тут всё корректно)
        all_ocr_text = ""
        for page_tokens in pages_data:
            h, t = ocr_to_grid_markdown(page_tokens, y_threshold=5)
            if t: all_ocr_text += f"\n\n{t}"
            
        return f"{header_md}\n\n{all_ocr_text.strip()}"
            
    except Exception as e:
        return f"Error building PDF markdown: {e}"


def excel_to_grid_markdown(file_path: str) -> str:
    """
    Converts an Excel invoice to Markdown using Pandas.
    Cleans up empty rows/cols and removes 'Unnamed' headers to avoid token explosion.
    """
    try:
        if file_path.lower().endswith('.csv'):
            df = pd.read_csv(file_path, dtype=str, sep=None, engine='python', on_bad_lines='skip')
        else:
            # Читаем без заголовков, чтобы не потерять первую строку полезных данных
            df = pd.read_excel(file_path, header=None, dtype=str)
    except Exception as e:
        return f"Error reading spreadsheet: {e}"

    # Drop fully empty cols and rows
    df = df.dropna(axis=1, how='all')
    df = df.dropna(axis=0, how='all')
    
    # Очищаем данные через карту (безопасно для имен колонок)
    df = df.map(clean_val)
    
    # Даем пустые имена колонкам (т.к. мы читали без заголовков, имена будут 0, 1, 2...)
    # Или заменяем Unnamed, если pandas все же их создал
    df.columns = ["" if (isinstance(c, int) or "Unnamed" in str(c)) else c for c in df.columns]
    
    return df.to_markdown(index=False)


def detect_pdf_type(file_path: str) -> str:
    """
    Detects if a PDF is text-based or a scan by checking the first page.
    """
    import pdfplumber
    try:
        with pdfplumber.open(file_path) as pdf:
            if not pdf.pages:
                return "SCAN_PDF"
            text = pdf.pages[0].extract_text() or ""
            if len(text.strip()) > 50:
                return "TEXT_PDF"
            return "SCAN_PDF"
    except Exception as e:
        print(f"PDF Type Detection error: {e}")
        return "SCAN_PDF"


def _load_client_settings() -> dict:
    """Load client (Buyer) config from client_settings.json next to this file."""
    import os, json
    cfg_path = os.path.join(os.path.dirname(__file__), "client_settings.json")
    try:
        with open(cfg_path, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {"inn": "5905271743", "name": "ММК-Пермь"}


def _discover_supplier(words: list, client_inn: str) -> dict:
    """
    DISCOVERY PASS — Phase 1.
    Scan all words to find any valid INN != client_inn.
    Then search Y ± 40px neighborhood for a legal-form org name.
    Returns: {"inn": str, "name": str}
    """
    import re
    ORG_PATTERN = re.compile(
        r'(ООО|ОАО|ЗАО|ПАО|АО|ИП|НКО|МУП|ГУП|ФГУП)\s*[«"\']?[\w\s\-–—]+[»"\'"]?',
        re.IGNORECASE
    )
    BANK_STOP = re.compile(r'(СБЕРБАНК|БАНК|БАНК\s+ПОЛУЧАТЕЛ|ВОЛГО|НОВГОРОД)', re.IGNORECASE)

    supplier_inn = None
    supplier_inn_y = None

    for w in words:
        txt = w['text']
        if len(txt) < 10:
            continue
        cand = validate_and_clean_inn(txt)
        if cand and cand != client_inn:
            supplier_inn = cand
            supplier_inn_y = w['y']
            break

    if not supplier_inn:
        return {"inn": None, "name": None}

    # Group words into text rows
    words_sorted = sorted(words, key=lambda w: w['y'])
    rows_grouped = []
    if words_sorted:
        cur_row = [words_sorted[0]]
        for i in range(1, len(words_sorted)):
            if abs(words_sorted[i]['y'] - cur_row[0]['y']) < 12:
                cur_row.append(words_sorted[i])
            else:
                rows_grouped.append(cur_row)
                cur_row = [words_sorted[i]]
        rows_grouped.append(cur_row)

    # Convert rows to text and find supplier INN row index
    row_texts = [" ".join(w['text'] for w in sorted(r, key=lambda x: x['x'])) for r in rows_grouped]
    target_idx = -1
    for i, r_txt in enumerate(row_texts):
        if supplier_inn in r_txt:
            target_idx = i
            break

    supplier_name = None
    if target_idx != -1:
        start_idx = max(0, target_idx - 2)
        end_idx = min(len(row_texts), target_idx + 3)
        for i in range(start_idx, end_idx):
            txt = row_texts[i]
            if BANK_STOP.search(txt):
                continue
            m = ORG_PATTERN.search(txt)
            if m:
                supplier_name = m.group(0).strip()
                break

    return {"inn": supplier_inn, "name": supplier_name}


def ocr_to_grid_markdown(words: list, y_threshold=15) -> tuple:
    """
    STAGE 3.2 — Spatial Baskets & Component Tokens (Basket Parsing).
    """
    if not words:
        return "", ""
    import re
    import math

    # ── Phase 0: Config & Constants ──────────────────────────────────────────
    cfg = _load_client_settings()
    CLIENT_INN  = cfg.get("client_inn",  "5905271743")
    CLIENT_NAME = cfg.get("name_keywords", ["ММК-Пермь"])[0] if cfg.get("name_keywords") else "ММК-Пермь"

    # ── Phase 0.1: Raw Stream Capture ────────────────────────────────────────
    # Capture original order before any sorting
    raw_stream_copy = list(words)

    # ── Phase 1: Classification & Table detection ────────────────────────────
    words.sort(key=lambda w: w['y'])
    rows_grouped = []
    if words:
        cur_row = [words[0]]
        for i in range(1, len(words)):
            if abs(words[i]['y'] - cur_row[0]['y']) < y_threshold: # Adaptive Threshold
                cur_row.append(words[i])
            else:
                rows_grouped.append(cur_row)
                cur_row = [words[i]]
        rows_grouped.append(cur_row)

    header_words_set = set() # To identify header words in raw stream
    state = "HEADER"
    for row in rows_grouped:
        row.sort(key=lambda x: x['x'])
        rl = " ".join(w['text'] for w in row).strip()
        l_line = rl.lower()

        if state == "HEADER" and any(k in l_line for k in ["наименование", "кол-во", "цена"]):
            state = "TABLE"
            continue
        
        if state == "HEADER":
            for w in row:
                header_words_set.add(id(w))
            
    all_header_words = [w for w in words if id(w) in header_words_set]

    # ── Phase 2: find_header_start_y (Universal Anchor) ─────────────────────
    def find_header_start_y(h_words: list) -> tuple:
        """
        Ищет верхнюю границу полезных данных.
        Возвращает (Y_start, Тип_якоря)
        """
        # Tags for Invoice
        tags = {"ИНН", "КПП", "БИК", "Сч. №", "Банк", "получателя"}
        
        words_sorted_y = sorted(h_words, key=lambda w: w['y'])
        
        # Priority 1: Tag Cloud (Invoice)
        for i, w in enumerate(words_sorted_y):
            if any(t in w['text'] for t in tags):
                # Count tags in 40px vertical window
                cloud_count = 1
                for j in range(i + 1, len(words_sorted_y)):
                    w2 = words_sorted_y[j]
                    if w2['y'] - w['y'] > 40: break
                    if any(t in w2['text'] for t in tags):
                        cloud_count += 1
                if cloud_count >= 2:
                    return w['y'], "Tag Cloud (Invoice)"

        # Priority 2: KP / Date marker (Clause)
        for w in h_words:
            t = w['text'].lower()
            if "коммерческое" in t or "предложение" in t or t == "от:":
                return w['y'], "KP / Date Marker"
                
        return 0.0, "None (Default)"

    # Identify Y_start
    y_start, anchor_type = find_header_start_y(all_header_words)
    
    # ── Phase 3: Slicing & Partitioning ──────────────────────────────────────
    # Crop words below anchor
    clean_words = sorted([w for w in all_header_words if w['y'] >= (y_start - 10)], key=lambda x: (x['y'], x['x']))
    
    # Group into lines (Y-Snapping)
    current_lines_data = [] # List of (y, text)
    if clean_words:
        row = [clean_words[0]]
        for i in range(1, len(clean_words)):
            if abs(clean_words[i]['y'] - row[0]['y']) < 15:
                row.append(clean_words[i])
            else:
                row.sort(key=lambda x: x['x'])
                current_lines_data.append((row[0]['y'], " ".join(w['text'] for w in row)))
                row = [clean_words[i]]
        row.sort(key=lambda x: x['x'])
        current_lines_data.append((row[0]['y'], " ".join(w['text'] for w in row)))

    # Find Y_title (Separator line)
    y_title = -1
    title_pattern = re.compile(r'(Счет|Счёт|Коммерческое предложение).*?№', re.IGNORECASE)
    for y, line in current_lines_data:
        if title_pattern.search(line):
            y_title = y
            break
            
    # Default Y_title if not found
    if y_title == -1:
        # Fallback to the middle of the header lines if title not found
        y_title = current_lines_data[len(current_lines_data)//2][0] if current_lines_data else 0

    # Partition into zones
    bank_lines = [l for y, l in current_lines_data if y < y_title]
    entities_lines = [l for y, l in current_lines_data if y >= y_title]
    
    bank_data_text = "\n".join(bank_lines)
    entities_data_text = "\n".join(entities_lines)

    # ── Phase 4: Final Output ─────────────────────────────────────────────
    header_md = "\n".join([
        "### [UNIVERSAL_DUMP] ###",
        f"- Found_Anchor: {anchor_type}",
        f"- Y_Start: {y_start}",
        f"- Y_Title: {y_title}",
        "",
        "### [ZONE_BANK_RAW] ###",
        bank_data_text if bank_data_text else "Пусто",
        "",
        "### [ZONE_ENTITIES_RAW] ###",
        entities_data_text if entities_data_text else "Пусто",
        "",
        "### [CLEAN_HEADER_LINES_CHECK] ###",
        "\n".join([f"{i}. {l}" for i, (y, l) in enumerate(current_lines_data)])
    ])

    return header_md, "### TABLE SUPPRESSED FOR TEST ###"

# --- КОНСТАНТЫ КЛИЕНТА (ММК-Пермь) ---
CLIENT_INN = "5905271743"
CLIENT_ADDRESS = "пермский край, г пермь, ул левченко, дом 1"

def extract_flat_tokens_from_yandex(ocr_json):
    """Превращает сложный JSON Яндекса в плоский список {'text': str, 'y_start': int, 'x_start': int}"""
    tokens = []
    # АГЕНТ: Адаптируй эту функцию под структуру OCR JSON в проекте
    # Пример обхода: pages -> blocks -> lines -> words -> boundingBox
    try:
        # Assuming OCR JSON is passed directly or is embedded in a specific structure
        # The logic here follows the standard Yandex Vision block format
        if isinstance(ocr_json, list) and len(ocr_json) > 0 and 'results' in ocr_json[0]:
            pages = ocr_json[0].get('results', [])[0].get('results', [])[0].get('textDetection', {}).get('pages', [])
        elif isinstance(ocr_json, dict) and 'pages' in ocr_json:
            pages = ocr_json.get('pages', [])
        else:
            # Maybe the input is just words list (as previously used)
            if isinstance(ocr_json, list) and len(ocr_json) > 0 and 'text' in ocr_json[0] and ('y' in ocr_json[0] or 'y_start' in ocr_json[0]):
                for w in ocr_json:
                    y = w.get('y', w.get('y_start', 0))
                    x = w.get('x', w.get('x_start', 0))
                    tokens.append({'text': w['text'], 'y_start': y, 'x_start': x})
                return tokens
            pages = []
            
        for page in pages:
            for block in page.get('blocks', []):
                for line in block.get('lines', []):
                    for word in line.get('words', []):
                        text = word.get('text', '')
                        vertices = word.get('boundingBox', {}).get('vertices', [])
                        if text and vertices:
                            y_start = min(int(v.get('y', 0)) for v in vertices)
                            x_start = min(int(v.get('x', 0)) for v in vertices)
                            tokens.append({'text': text, 'y_start': y_start, 'x_start': x_start})
    except Exception as e:
        print(f"Error parsing tokens: {e}")
    return tokens

def normalize_to_standard_grid(words, width, height):
    """Приводит любые координаты (points или pixels) к единой сетке 0-1000."""
    standard_words = []
    for w in words:
        w_width = width if width else 1000
        w_height = height if height else 1000
        x = w.get('x', w.get('x_start', 0))
        y = w.get('y', w.get('y_start', 0))
        standard_words.append({
            "text": w["text"],
            "x": (x / w_width) * 1000,
            "y": (y / w_height) * 1000,
        })
    return standard_words

def build_zonal_markdown(words, is_digital=False):
    """Сборщик зон. Разделяет логику для сканов (с якорем) и цифровых PDF (без якоря)."""
    if not words: return "NO_TEXT_FOUND"

    # 1. Сортируем слова сверху вниз, слева направо
    words.sort(key=lambda x: (x['y'], x['x']))

    # 2. Склеиваем слова в строки с учетом горизонтальных пробелов
    lines_data = []
    if words:
        current_line_words = [words[0]]
        last_y = words[0]['y']
        # Чувствительность по Y: для цифры строже, для скана мягче
        y_thresh = 5 if is_digital else 15

        for w in words[1:]:
            if abs(w['y'] - last_y) < y_thresh:
                current_line_words.append(w)
            else:
                # Строка собрана, склеиваем слова
                current_line_words.sort(key=lambda x: x['x'])
                line_text_parts = []
                for i, cw in enumerate(current_line_words):
                    # Если между словами большая дыра (> 50 единиц) - ставим 4 пробела (разделитель колонок)
                    if i > 0 and (cw['x'] - current_line_words[i-1]['x']) > 50:
                        line_text_parts.append("    " + cw['text'])
                    else:
                        line_text_parts.append(cw['text'])
                
                line_text = " ".join(line_text_parts).replace("     ", "    ").strip()
                lines_data.append({"text": line_text, "y": last_y})

                current_line_words = [w]
                last_y = w['y']

        # Добиваем последнюю строку
        current_line_words.sort(key=lambda x: x['x'])
        line_text_parts = []
        for i, cw in enumerate(current_line_words):
            if i > 0 and (cw['x'] - current_line_words[i-1]['x']) > 50:
                line_text_parts.append("    " + cw['text'])
            else:
                line_text_parts.append(cw['text'])
        line_text = " ".join(line_text_parts).replace("     ", "    ").strip()
        lines_data.append({"text": line_text, "y": last_y})

    # =======================================================
    # 3. ФОРМИРОВАНИЕ ЗОН (РАЗВИЛКА ЛОГИКИ)
    # =======================================================

    if is_digital:
        # --- ПУТЬ ДЛЯ ЦИФРОВЫХ PDF ---
        # Умный ограничитель шапки: берем верхнюю часть, но останавливаемся перед таблицей
        header_lines = []
        for l in lines_data:
            if l['y'] >= 650:
                break
            
            txt_lower = l['text'].lower()
            
            # Ищем признаки заголовка таблицы (пересечение сущностей и метрик)
            # Мы ищем И номер/название И одновременно количество/цену/сумму
            has_identity = any(m in txt_lower for m in ["№", "наименование", "товар", "услуги", "артикул"])
            has_metrics = any(m in txt_lower for m in ["кол-во", "количество", "цена", "сумма", "ед.изм", "ед. изм"])
            
            if has_identity and has_metrics:
                # Мы дошли до шапки таблицы — прекращаем сбор реквизитов
                break
                
            header_lines.append(l['text'])

        output = "### [ZONE_ENTITIES_RAW] ###\n"
        output += "\n".join(header_lines)
        return output

    else:
        # --- ПУТЬ ДЛЯ СКАНОВ (ОСТАВЛЯЕМ СТАРУЮ ЛОГИКУ БЕЗ ИЗМЕНЕНИЙ) ---
        # Ищем якорь "Счет №"
        y_anchor = 400
        for line in lines_data:
            txt = line['text'].lower()
            if "счет" in txt and ("№" in txt or "на" in txt or "оплату" in txt or "n " in txt):
                y_anchor = line['y']
                break

        bank_zone = [l['text'] for l in lines_data if l['y'] < y_anchor - 5]
        entities_zone = [l['text'] for l in lines_data if y_anchor - 5 <= l['y'] < y_anchor + 350]

        output = "### [ZONE_BANK_RAW] ###\n"
        output += "\n".join(bank_zone)
        output += "\n\n### [ZONE_ENTITIES_RAW] ###\n"
        output += "\n".join(entities_zone)

        return output

def clean_and_build_markdown(ocr_json_or_words):
    """Единый адаптер (маршрутизатор) для сканов и PDF."""
    
    # Если это список словарей с ключом 'text' (из PDF pdfplumber) -> ЭТО ЦИФРОВОЙ PDF
    if isinstance(ocr_json_or_words, list) and len(ocr_json_or_words) > 0 and isinstance(ocr_json_or_words[0], dict) and 'text' in ocr_json_or_words[0]:
        # Передаем is_digital=True и нормализуем под стандартный размер A4 в пунктах
        return build_zonal_markdown(normalize_to_standard_grid(ocr_json_or_words, 595, 842), is_digital=True)

    # Иначе парсим как ответ Yandex Vision -> ЭТО СКАН
    tokens = extract_flat_tokens_from_yandex(ocr_json_or_words)
    if not tokens: return "NO_TEXT_FOUND"
    
    # Передаем is_digital=False
    return build_zonal_markdown(normalize_to_standard_grid(tokens, 1650, 2330), is_digital=False)

def generate_invoice_summary(data: dict) -> str:
    """
    Sprint 1/2: Генерация Markdown-сводки для правой панели.
    Извлекает ключевые поля из структуры документа LLM.
    """
    doc = data.get("document", {})
    
    # Извлекаем основные поля
    inv_no = doc.get("invoice_number", "---")
    inv_date = doc.get("invoice_date", "---")
    supplier = doc.get("organization_name", "---")
    inn = doc.get("inn", "---")
    total = doc.get("total_amount", "---")
    currency = doc.get("currency", "RUB")
    
    # Формируем Markdown таблицу
    lines = [
        "### 🧾 Сводка по документу",
        "",
        "| Поле | Значение |",
        "| :--- | :--- |",
        f"| **Номер счета** | {inv_no} |",
        f"| **Дата** | {inv_date} |",
        f"| **Поставщик** | {supplier} |",
        f"| **ИНН** | {inn} |",
        f"| **Сумма Итого** | **{total} {currency}** |",
        "",
        "> [!NOTE]",
        "> Результаты извлечены нейросетью. Пожалуйста, проверьте точность данных в панели ниже."
    ]
    
    return "\n".join(lines)


def extract_digital_words_as_ocr(pdf_path: str) -> list:
    """
    Extracts text bounding boxes from a digital PDF using pdfplumber.
    Returns: A list of lists (one per page), where each page is a list of word dicts:
             {'text': str, 'x': float, 'y': float, 'w': float, 'h': float}
    Coordinates are normalized to 0-1000 standard grid internally.
    Deduplicates double-extracted tokens (common in some PDFs).
    """
    import pdfplumber

    pages_data = []
    
    with pdfplumber.open(pdf_path) as pdf:
        for pg in pdf.pages:
            words = pg.extract_words(x_tolerance=2, y_tolerance=3, keep_blank_chars=False, use_text_flow=True)
            
            p_width = pg.width or 595
            p_height = pg.height or 842
            
            page_words = []
            seen_tokens = set() # To prevent duplicates like (123, 123)
            
            for w in words:
                text = w["text"].strip()
                if not text: continue
                
                x0, top, x1, bottom = w["x0"], w["top"], w["x1"], w["bottom"]
                
                # Возвращаем ПРЯМЫЕ координаты (points), нормализация будет в clean_and_build_markdown
                nx = x0
                ny = top
                nw = (x1 - x0)
                nh = (bottom - top)
                
                # Fingerprint to ignore duplicates (используем округление до целых пунктов)
                fpr = (text, round(nx), round(ny)) 
                if fpr in seen_tokens:
                    continue
                seen_tokens.add(fpr)
                
                page_words.append({
                    "text": text,
                    "x": nx,
                    "y": ny,
                    "w": nw,
                    "h": nh
                })
            pages_data.append(page_words)
            
    return pages_data


# ============================================================================
# PDF TABLE EXTRACTION FUNCTIONS (Этап 1: Извлечение позиций из цифровых PDF)
# ============================================================================

def group_words_into_lines(words: list, y_threshold: int = 5) -> list:
    """
    Группирует слова в строки по Y-координате.
    
    Args:
        words: Список словарей с ключами 'text', 'x', 'y', 'w', 'h'
        y_threshold: Порог для объединения в одну строку (по умолчанию 5)
    
    Returns:
        Список кортежей (y_coordinate, line_text)
    
    Example:
        >>> words = [
        ...     {"text": "ООО", "x": 50, "y": 100, "w": 30, "h": 12},
        ...     {"text": "Поставщик", "x": 90, "y": 100, "w": 80, "h": 12},
        ...     {"text": "Счет", "x": 50, "y": 130, "w": 40, "h": 12},
        ... ]
        >>> group_words_into_lines(words)
        [(100, "ООО Поставщик"), (130, "Счет")]
    """
    if not words:
        return []
    
    words_sorted = sorted(words, key=lambda w: w['y'])
    lines = []
    current_line = [words_sorted[0]]
    
    for word in words_sorted[1:]:
        if abs(word['y'] - current_line[0]['y']) < y_threshold:
            current_line.append(word)
        else:
            # Сортируем по X и склеиваем
            current_line.sort(key=lambda w: w['x'])
            line_text = ' '.join(w['text'] for w in current_line)
            lines.append((current_line[0]['y'], line_text))
            current_line = [word]
    
    # Последняя строка
    if current_line:
        current_line.sort(key=lambda w: w['x'])
        line_text = ' '.join(w['text'] for w in current_line)
        lines.append((current_line[0]['y'], line_text))
    
    return lines


def detect_table_columns(words: list, x_threshold: int = 20) -> list:
    """
    Определяет границы колонок таблицы по X-координатам слов.
    
    Args:
        words: Список словарей с ключами 'text', 'x', 'y', 'w', 'h'
        x_threshold: Порог для объединения в одну колонку (по умолчанию 20)
    
    Returns:
        Список средних X-координат колонок (отсортированный)
    
    Example:
        >>> words = [
        ...     {"text": "№", "x": 50, "y": 100, ...},
        ...     {"text": "Наименование", "x": 100, "y": 100, ...},
        ...     {"text": "Цена", "x": 350, "y": 100, ...},
        ... ]
        >>> detect_table_columns(words)
        [50, 100, 350]
    """
    if not words:
        return []
    
    x_coords = sorted(set(w['x'] for w in words))
    
    if not x_coords:
        return []
    
    columns = []
    current_col = [x_coords[0]]
    
    for x in x_coords[1:]:
        if x - current_col[-1] < x_threshold:
            current_col.append(x)
        else:
            columns.append(sum(current_col) / len(current_col))
            current_col = [x]
    
    if current_col:
        columns.append(sum(current_col) / len(current_col))
    
    return columns


def words_to_markdown_grid(words: list, y_threshold: int = 5, x_threshold: int = 20) -> str:
    """
    Конвертирует координатные слова в Markdown-таблицу.
    
    Args:
        words: Список словарей с ключами 'text', 'x', 'y', 'w', 'h'
        y_threshold: Порог для группировки строк
        x_threshold: Порог для определения колонок
    
    Returns:
        Markdown-таблица в виде строки
    
    Example:
        >>> words = [
        ...     {"text": "№", "x": 50, "y": 100, ...},
        ...     {"text": "Наименование", "x": 100, "y": 100, ...},
        ...     {"text": "1", "x": 52, "y": 120, ...},
        ...     {"text": "Товар А", "x": 105, "y": 120, ...},
        ... ]
        >>> words_to_markdown_grid(words)
        '| № | Наименование |\\n| 1 | Товар А |'
    """
    if not words:
        return ""
    
    # 1. Определяем колонки
    columns = detect_table_columns(words, x_threshold)
    
    if not columns:
        return ""
    
    # 2. Группируем в строки
    lines = group_words_into_lines(words, y_threshold)
    
    # 3. Распределяем слова по колонкам
    grid = []
    
    for line_y, _ in lines:
        line_words = [w for w in words if abs(w['y'] - line_y) < y_threshold]
        line_words.sort(key=lambda w: w['x'])
        
        row = [''] * len(columns)
        
        for word in line_words:
            # Находим ближайшую колонку
            col_idx = min(range(len(columns)),
                         key=lambda i: abs(columns[i] - word['x']))
            
            # Добавляем текст в ячейку
            if row[col_idx]:
                row[col_idx] += ' ' + word['text']
            else:
                row[col_idx] = word['text']
        
        grid.append(row)
    
    # 4. Очищаем и объединяем колонки
    if not grid:
        return ""
    
    cleaned_grid = clean_and_merge_table_columns(grid)
    
    # 5. Конвертируем в Markdown
    md_lines = []
    for row in cleaned_grid:
        md_lines.append('| ' + ' | '.join(row) + ' |')
    
    return '\n'.join(md_lines)


def clean_and_merge_table_columns(grid: list) -> list:
    """
    Очищает и объединяет колонки таблицы для улучшения парсинга GPT.
    
    Выполняет:
    1. Удаление пустых колонок (>50% пустых ячеек)
    2. Объединение соседних колонок с короткими значениями (только единицы измерения)
    3. Упрощение заголовка (удаление дублирующих строк)
    
    Args:
        grid: Двумерный массив строк таблицы
    
    Returns:
        Очищенный grid
    """
    if not grid or len(grid) < 2:
        return grid
    
    # Шаг 1: Удаляем полностью пустые колонки
    num_cols = len(grid[0])
    cols_to_keep = []
    
    for col_idx in range(num_cols):
        non_empty_count = sum(1 for row in grid if col_idx < len(row) and row[col_idx].strip())
        # Колонка должна быть заполнена хотя бы наполовину
        if non_empty_count >= len(grid) * 0.5:
            cols_to_keep.append(col_idx)
    
    # Применяем фильтр колонок
    filtered_grid = []
    for row in grid:
        filtered_row = [row[i] if i < len(row) else '' for i in cols_to_keep]
        filtered_grid.append(filtered_row)
    
    if not filtered_grid:
        return grid
    
    # Шаг 2: Объединяем ТОЛЬКО единицы измерения с количеством
    # Например: "1" + "шт" → "1 шт", "46" + "шт" → "46 шт"
    # НЕ объединяем другие колонки!
    merged_grid = []
    
    for row in filtered_grid:
        merged_row = []
        i = 0
        while i < len(row):
            cell = row[i].strip()
            
            # Проверяем, нужно ли объединить со следующей ячейкой
            if i + 1 < len(row):
                next_cell = row[i + 1].strip()
                
                # ТОЛЬКО случай: Число + единица измерения (шт, м, кг и т.д.)
                # Единица измерения должна быть короткой (< 10 символов) и не содержать цифр
                should_merge = False
                
                if cell and next_cell:
                    # Текущая ячейка - число, следующая - единица измерения
                    if (cell.replace(',', '').replace('.', '').replace(' ', '').isdigit() and
                        len(next_cell) < 10 and
                        not any(c.isdigit() for c in next_cell)):
                        should_merge = True
                
                if should_merge:
                    merged_row.append(f"{cell} {next_cell}")
                    i += 2  # Пропускаем следующую ячейку
                    continue
            
            merged_row.append(cell)
            i += 1
        
        merged_grid.append(merged_row)
    
    # Шаг 3: Упрощаем заголовок (удаляем вторую строку, если она дублирует первую)
    if len(merged_grid) >= 2:
        header1 = merged_grid[0]
        header2 = merged_grid[1]
        
        # Проверяем, является ли вторая строка продолжением заголовка
        is_header_continuation = True
        for cell in header2:
            cell_lower = cell.lower().strip()
            # Если во второй строке есть числа или длинный текст, это не заголовок
            if cell_lower and (any(c.isdigit() for c in cell) or len(cell) > 20):
                is_header_continuation = False
                break
        
        # Если вторая строка - продолжение заголовка, объединяем или удаляем
        if is_header_continuation:
            # Объединяем заголовки
            combined_header = []
            for i in range(max(len(header1), len(header2))):
                h1 = header1[i] if i < len(header1) else ''
                h2 = header2[i] if i < len(header2) else ''
                
                if h1 and h2:
                    combined_header.append(f"{h1} {h2}".strip())
                else:
                    combined_header.append((h1 or h2).strip())
            
            merged_grid[0] = combined_header
            merged_grid.pop(1)  # Удаляем вторую строку заголовка
    
    return merged_grid


def is_subtotal_line(lines: list, current_idx: int, words: list = None) -> bool:
    """
    Проверяет, является ли строка промежуточным итогом (Peak-ahead).
    
    Логика: Если после строки с "итого" есть строки с якорями (короткие значения
    в начале строки, например номера позиций), то это промежуточный итог.
    
    Args:
        lines: Список кортежей (y_coordinate, line_text)
        current_idx: Индекс текущей строки
        words: Опциональный список слов для более точной проверки
    
    Returns:
        True если это промежуточный итог, False если финальный
    
    Example:
        >>> lines = [
        ...     (100, "1 Товар А 10 100"),
        ...     (120, "Итого по разделу 100"),
        ...     (140, "2 Товар Б 5 50"),  # <- есть якорь (2)
        ... ]
        >>> is_subtotal_line(lines, 1)
        True  # Это промежуточный итог
    """
    # Проверяем следующие 3 строки
    for look_idx in range(current_idx + 1, min(current_idx + 4, len(lines))):
        _, line_text = lines[look_idx]
        
        # Ищем якорь: короткое значение в начале строки
        tokens = line_text.strip().split()
        if tokens:
            first_token = tokens[0]
            # Якорь: короткое значение (< 15 символов) и не "итого"
            if len(first_token) < 15 and first_token.lower() != 'итого':
                return True  # Промежуточный итог
    
    return False  # Финальный итог


def extract_pdf_table_region(pages_data: list) -> str:
    """
    Извлекает только табличную часть из цифрового PDF.
    
    Алгоритм:
    1. Ищет начало таблицы по маркерам (наименование, кол-во, цена, сумма)
    2. Ищет конец таблицы по стоп-словам (итого, всего к оплате) с Peak-ahead
    3. Извлекает слова из найденного региона
    4. Конвертирует в Markdown-таблицу
    
    Args:
        pages_data: Список страниц, каждая страница - список словарей слов
                   [{'text': str, 'x': float, 'y': float, 'w': float, 'h': float}, ...]
    
    Returns:
        Markdown-таблица в виде строки
    
    Example:
        >>> pages_data = [[
        ...     {"text": "Наименование", "x": 100, "y": 300, ...},
        ...     {"text": "Цена", "x": 350, "y": 300, ...},
        ...     {"text": "Товар А", "x": 100, "y": 320, ...},
        ...     {"text": "100", "x": 355, "y": 320, ...},
        ... ]]
        >>> extract_pdf_table_region(pages_data)
        '| Наименование | Цена |\\n| Товар А | 100 |'
    """
    if not pages_data:
        return ""
    
    # Маркеры начала и конца таблицы
    table_markers = ['наименование', 'кол-во', 'цена', 'сумма', 'количество', 'товар', 'артикул']
    stop_words = ['итого', 'всего к оплате', 'всего наименований', 'руководитель', 'м.п.', 'главный бухгалтер']
    
    # 1. ПОИСК НАЧАЛА ТАБЛИЦЫ
    table_start_y = None
    table_start_page = 0
    
    for page_idx, page_words in enumerate(pages_data):
        if not page_words:
            continue
            
        lines = group_words_into_lines(page_words, y_threshold=5)
        
        for line_y, line_text in lines:
            line_lower = line_text.lower()
            matches = sum(1 for marker in table_markers if marker in line_lower)
            
            if matches >= 2:
                table_start_y = line_y
                table_start_page = page_idx
                break
        
        if table_start_y is not None:
            break
    
    if table_start_y is None:
        return ""  # Таблица не найдена
    
    # 2. ПОИСК КОНЦА ТАБЛИЦЫ
    table_end_y = None
    table_end_page = len(pages_data) - 1
    
    for page_idx in range(table_start_page, len(pages_data)):
        if not pages_data[page_idx]:
            continue
            
        lines = group_words_into_lines(pages_data[page_idx], y_threshold=5)
        
        for idx, (line_y, line_text) in enumerate(lines):
            line_lower = line_text.lower()
            
            if any(stop in line_lower for stop in stop_words):
                # Peak-ahead: проверяем промежуточный итог
                is_subtotal = is_subtotal_line(lines, idx, pages_data[page_idx])
                
                # Дополнительная проверка: если на текущей странице нет якорей,
                # проверяем следующую страницу на наличие маркеров таблицы
                if not is_subtotal and page_idx + 1 < len(pages_data):
                    next_page_lines = group_words_into_lines(pages_data[page_idx + 1], y_threshold=5)
                    
                    # Ищем маркеры таблицы на следующей странице
                    for next_line_y, next_line_text in next_page_lines[:10]:  # Проверяем первые 10 строк
                        next_line_lower = next_line_text.lower()
                        matches = sum(1 for marker in table_markers if marker in next_line_lower)
                        
                        # Если нашли маркеры таблицы - это промежуточный итог
                        if matches >= 2:
                            is_subtotal = True
                            break
                
                if not is_subtotal:
                    table_end_y = line_y
                    table_end_page = page_idx
                    break
        
        if table_end_y is not None:
            break
    
    # Если конец не найден, берем до конца документа
    if table_end_y is None:
        table_end_y = float('inf')
        table_end_page = len(pages_data) - 1
    
    # 3. ИЗВЛЕЧЕНИЕ СЛОВ ИЗ РЕГИОНА
    table_words = []
    
    for page_idx in range(table_start_page, table_end_page + 1):
        for word in pages_data[page_idx]:
            # Фильтруем по Y-координатам
            if page_idx == table_start_page and word['y'] < table_start_y:
                continue
            if page_idx == table_end_page and word['y'] > table_end_y:
                continue
            
            table_words.append(word)
    
    # 4. КОНВЕРТАЦИЯ В MARKDOWN
    markdown_table = words_to_markdown_grid(table_words, y_threshold=5, x_threshold=35)
    
    return markdown_table


def excel_to_markdown_header(file_path: str, file_extension: str) -> str:
    """
    Универсальный препроцессор. Читает Excel/CSV строго как текст, 
    лечит кавычки, отрезает таблицу товаров и возвращает Markdown.
    """
    try:
        ext = file_extension.lower()
        if ext == '.csv':
            df = pd.read_csv(file_path, dtype=str, sep=None, engine='python', on_bad_lines='skip')
        elif ext in ['.xlsx', '.xls']:
            df = pd.read_excel(file_path, header=None, dtype=str)
        else:
            return ""
    except Exception as e:
        print(f"Error reading spreadsheet: {e}")
        return ""

    # Primary cleanup
    df = df.dropna(axis=1, how='all')
    
    # Bulk cleaning
    df = df.map(clean_val)

    # 3. Ищем начало таблицы товаров (Smart Cut-off)
    table_start_idx = len(df)
    target_words = {'кол-во', 'цена', 'количество', 'сумма', 'товар', 'наименование', 'артикул'}
    
    for idx, row in df.iterrows():
        row_text = ' '.join(map(str, row.values)).lower()
        matches = sum(1 for word in target_words if word in row_text)
        if matches >= 2:
            table_start_idx = idx
            break
            
    # Поиск конца таблицы для корректного отсечения подвала
    table_end_idx = table_start_idx
    stop_footer_words = {'итого', 'всего к оплате', 'всего наименований', 'сумма ндс', 'в т.ч. ндс', 'итого с ндс'}
    
    if table_start_idx < len(df):
        for idx in range(table_start_idx + 1, len(df)):
            row_text = ' '.join(map(str, df.iloc[idx].values)).lower()
            if any(stop in row_text for stop in stop_footer_words):
                table_end_idx = idx
                break
        if table_end_idx == table_start_idx:
            # Если стоп-слова не найдены, считаем что таблица до конца (или была ошибка поиска)
            table_end_idx = len(df)
            
    # Assembly
    header_df = df.iloc[:table_start_idx].head(60)
    header_df.columns = ["" if (isinstance(c, int) or "Unnamed" in str(c)) else c for c in header_df.columns]
    
    footer_text = ""
    if table_end_idx < len(df):
        footer_df = df.iloc[table_end_idx:]
        if not footer_df.empty:
            footer_df = footer_df.dropna(axis=1, how='all').dropna(axis=0, how='all')
            if not footer_df.empty:
                footer_df.columns = ["" if (isinstance(c, int) or "Unnamed" in str(c)) else c for c in footer_df.columns]
                footer_text = footer_df.to_markdown(index=False)

    markdown_header = header_df.to_markdown(index=False) if not header_df.empty else ""
    
    if not markdown_header:
        return ""
        
    if footer_text:
        return f"{markdown_header}\n\n... [ТАБЛИЦА ТОВАРОВ СКРЫТА] ...\n\n### ПОДВАЛ ДОКУМЕНТА:\n{footer_text}"
    else:
        return markdown_header

def split_excel_to_md_sections(file_path: str) -> tuple[str, str, str]:
    """
    Разделяет Excel-документ на три MD-блока (header_md, items_md, footer_md).
    Глобальный сканер: ищет ВСЕ таблицы на ВСЕХ листах (до 10),
    затем выбирает «Абсолютного Лидера» — таблицу с наибольшим количеством строк.
    """
    try:
        if file_path.lower().endswith('.csv'):
            sheets_dict = {'Sheet1': pd.read_csv(file_path, dtype=str, sep=None, engine='python', on_bad_lines='skip')}
        else:
            sheets_dict = pd.read_excel(file_path, sheet_name=None, header=None, dtype=str)
    except Exception as e:
        print(f"Error reading spreadsheet: {e}")
        return "", "", ""

    target_words = {'кол-во', 'цена', 'количество', 'сумма', 'товар', 'наименование', 'артикул'}
    stop_words = {'итого', 'всего к оплате', 'всего наименований', 'сумма ндс', 'в т.ч. ндс', 'итого с ндс', 'всего'}
    summary_noise_words = {'итого', 'всего', 'ндс', 'сумм', 'рубл', 'копе'}

    candidates = []

    # Сканируем до 10 листов
    for sheet_name in list(sheets_dict.keys())[:10]:
        df = sheets_dict[sheet_name]
        df = df.map(clean_val)
        n_rows = len(df)

        # Контекст над первой таблицей (для бонуса «Спецификация»)
        sheet_name_lower = str(sheet_name).lower()

        # --- Поиск ВСЕХ таблиц на этом листе ---
        scan_from = 0
        while scan_from < n_rows:
            # 1. Поиск следующего заголовка таблицы
            table_start_idx = None
            for idx in range(scan_from, n_rows):
                row_text = ' '.join(map(str, df.iloc[idx].values)).lower()
                matches = sum(1 for word in target_words if word in row_text)
                if matches >= 2:
                    table_start_idx = idx
                    break

            if table_start_idx is None:
                break  # На этом листе больше нет таблиц

            # 2. Поиск конца таблицы (с Peak-ahead в 3 строки)
            table_end_idx = n_rows
            for idx in range(table_start_idx + 1, n_rows):
                row_text = ' '.join(map(str, df.iloc[idx].values)).lower()
                if any(stop in row_text for stop in stop_words):
                    # Peak-ahead: проверяем следующие 3 строки на наличие якоря
                    is_subtotal = False
                    for l_idx in range(idx + 1, min(idx + 4, n_rows)):
                        val_0 = str(df.iloc[l_idx, 0]).strip() if df.shape[1] > 0 else ""
                        val_1 = str(df.iloc[l_idx, 1]).strip() if df.shape[1] > 1 else ""
                        if (val_0 and len(val_0) < 15 and val_0.lower() != 'итого') or \
                           (val_1 and len(val_1) < 15 and val_1.lower() != 'итого'):
                            is_subtotal = True
                            break
                    if is_subtotal:
                        continue  # Промежуточный итог — продолжаем
                    table_end_idx = idx
                    break

            # Подсчет строк
            rows_count = table_end_idx - table_start_idx
            if rows_count <= 1:
                scan_from = table_end_idx + 1
                continue

            score = rows_count
            
            # Бонус за слово «Спецификация» в шапке листа или имени листа
            header_text = " ".join(
                " ".join(map(str, df.iloc[r].values))
                for r in range(0, table_start_idx)
            ).lower()
            if "спецификация" in header_text or "спецификация" in sheet_name_lower:
                score += 50

            # 3. Поиск подвала (пропускаем мусор сумм)
            real_footer_idx = n_rows
            if table_end_idx < n_rows:
                for idx in range(table_end_idx, n_rows):
                    row_text = ' '.join(map(str, df.iloc[idx].values)).lower()
                    if not row_text.strip():
                        continue
                    if any(noise in row_text for noise in summary_noise_words):
                        continue
                    real_footer_idx = idx
                    break

            candidates.append({
                'score': score,
                'rows': rows_count,
                'df': df,
                'start': table_start_idx,
                'end': table_end_idx,
                'footer': real_footer_idx,
                'sheet': sheet_name
            })

            # Продолжаем сканирование ПОСЛЕ конца этой таблицы
            scan_from = table_end_idx + 1

    if not candidates:
        return "", "", ""

    # Выбираем Абсолютного Лидера: таблица с максимальным score
    leader = max(candidates, key=lambda x: x['score'])
    df = leader['df']
    table_start_idx = leader['start']
    table_end_idx = leader['end']
    real_footer_idx = leader['footer']

    # Срезка блоков лидера
    header_df = df.iloc[:table_start_idx].dropna(axis=1, how='all').dropna(axis=0, how='all')
    items_df = df.iloc[table_start_idx:table_end_idx + 1].dropna(axis=1, how='all').dropna(axis=0, how='all')
    footer_df = df.iloc[real_footer_idx:].dropna(axis=1, how='all').dropna(axis=0, how='all')

    for d in [header_df, items_df, footer_df]:
        if not d.empty:
            d.columns = ["" if (isinstance(c, int) or "Unnamed" in str(c)) else c for c in d.columns]

    header_md = header_df.to_markdown(index=False) if not header_df.empty else ""
    items_md = items_df.to_markdown(index=False) if not items_df.empty else ""
    footer_md = footer_df.to_markdown(index=False) if not footer_df.empty else ""

    return header_md, items_md, footer_md






